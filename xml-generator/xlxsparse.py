import os
import xml.etree.ElementTree as ET
from xml.etree.ElementTree import Element, SubElement, Comment, tostring
# https://docs.python.org/2/library/xml.etree.elementtree.html
# https://docs.python.org/3/library/xml.etree.elementtree.html
from xml.dom.minidom import parse, parseString
# https://docs.python.org/3/library/xml.dom.minidom.html
from sys import platform
import time
import datetime
import operator
from openpyxl import Workbook
from openpyxl import load_workbook
from objxls import *
from itertools import islice
from xmlwrite import *
import sys
reload(sys)
sys.setdefaultencoding('utf-8')

def xlxsparse(objCourse, objSection, objModule):
    # objxlCourse = []
    # objxlSection = []
    # objxlCourse = []

    option = 0
    print "\n1. Import Excel File of Modification to a Moodle Backup at " + objCourse.location
    print "2. Write Excel File from Moodle Backup at " + objCourse.location
    print "3. Add Resource (COMING SOON!)"
    print "Any other value to abort"
    try:
        option = raw_input("(1 or 2)\n")
    except:
        option = input("(1 or 2\n)")

    def choice(option):
        if option == "1":
            readXL()
        elif option == "2":
            writeXL()
        else:
            print "Aborting"
            return 0
        return 1

    def readXL():
        """
        load_workbook and write the data from the excel workbook to three list objects
        for comparison with the list objects from the Moodle Backup
        """

        try:
            directory = raw_input("Please enter the full path of the XLSX document \n (e.g. /Users/milesexner/Desktop/Moodle-Course/xml-generator/WS800.01.2017S.SD.xlsx) : ")
        except:
            directory = input("Please enter the full path of the XLSX document \n (e.g. /Users/milesexner/Desktop/Moodle-Course/xml-generator/WS800.01.2017S.SD.xlsx) : ")

        # default
        if directory == "":
            directory = "/Users/milesexner/Desktop/Moodle-Course/xml-generator/WS800.01.2017S.SD.xlsx"

        # fullPath = os.path.join(directory, filename)
        fullPath = directory

        try:
            print "Path: " + fullPath
            wb = load_workbook(filename = fullPath)
            print "Workbook successfully loaded"

        except:
            print "Sorry, could not open %s. Please try again." % (fullPath)
            return 0

        importXLS(wb)
        # writeXML(objxlCourse, objxlSection, objxlCourse)
        return 1

    def importXLS(wb):
        """
        Grab the location of the Moodle_backup.xml to pass to impActivities, so it can write activity field changes to that file
        """
        location = impCourse(wb)
        numSections = impSections(wb)
        impActivities(wb, numSections, location)

        # wb.sheet
        # for sheet in wb:

        return numSections

    def impCourse(wb):
        ws = wb.get_sheet_by_name("Course")

        objxlCourse = xlcourse()
        objxlCourse.shortName = ws['B1'].value
        objxlCourse.fullName = ws['B2'].value
        objxlCourse.startDate = ws['B3'].value
        objxlCourse.location = ws['B4'].value



        """
        Test
        """
        print objxlCourse.shortName
        print objxlCourse.fullName
        print objxlCourse.startDate
        print objxlCourse.location


        writeCourse(objxlCourse)
        return objxlCourse.location

    def impSections(wb):
        ws = wb.get_sheet_by_name("Sections")
        objxlSection = []
        c = 1
        for col in islice(ws.columns, 1, None):
            c += 1
            # Add 1, since row 0 does not exist, step by number of rows
            for x in range(1, len(col)+1, 6):
                objxlSection.append(xlsection(ws.cell(row=x, column=c).value, ws.cell(row=x+1, column=c).value, ws.cell(row=x+2, column=c).value,
                ws.cell(row=x+3, column=c).value, ws.cell(row=x+4, column=c).value))
                # Test
                """
                print "Number: " + objxlSection[c-2].number
                print "Name: " + str(objxlSection[c-2].name)
                print "Summary: " + objxlSection[c-2].summary
                print "Location: " + objxlSection[c-2].location
                """
            writeSection(objxlSection[c-2])

        return c-1 # Number of Sections

    def impActivities(wb, numSections, location):
        objxlModule = []

        print "Number of Sections: " + str(numSections)
        b = 1
        for i in range(0, numSections):
            ws = wb.get_sheet_by_name("Section" + str(i))
            c = 1
            for col in islice(ws.columns, 1, None):
                c += 1
                b += 1
                # Add 1, since row 0 does not exist, step by number of rows
                numRows = 13
                for x in range(1, len(col)+1, numRows):
                    # Module/Activity Name (Type), Module/Activity ID, Location, Name, Intro, Content, URL, Grade, Due Date
                    objxlModule.append(xlmodule(ws.cell(row=x, column=c).value, ws.cell(row=x+1, column=c).value, ws.cell(row=x+2, column=c).value, ws.cell(row=x+3, column=c).value, ws.cell(row=x+4, column=c).value,
                    ws.cell(row=x+5, column=c).value, ws.cell(row=x+6, column=c).value, ws.cell(row=x+7, column=c).value,
                    ws.cell(row=x+8, column=c).value, ws.cell(row=x+9, column=c).value, ws.cell(row=x+10, column=c).value, ws.cell(row=x+11, column=c).value, ws.cell(row=x+11, column=c).value))



                    print "\nModule Type: " + str(objxlModule[b-2].modulename)
                    print "Module ID: " + str(objxlModule[b-2].moduleID)
                    print "Location: " + str(objxlModule[b-2].location)
                    print "Name: " + str(objxlModule[b-2].name)
                    print "Intro: " + str(objxlModule[b-2].intro)
                    print "Content: " + str(objxlModule[b-2].content)
                    print "URL: " + str(objxlModule[b-2].url)
                    print "Grade: " + str(objxlModule[b-2].grade)
                    print "Due Date: " + str(objxlModule[b-2].dueDate)


                writeActivity(objxlModule[c-2], location, c-2)

        return 1

    def writeXL():
        wb = Workbook()
        destFilename = objCourse.shortName + ".xlsx"
        print "File will be written as " + destFilename

        courseSheet(wb)
        sectionSheet(wb)
        activitySheet(wb)

        try:
            wb.save(filename = destFilename)
            print "File saved as: " + destFilename
        except:
            print "Save failed"

    def courseSheet(wb):
        wsCourse = wb.active
        wsCourse.title = "Course"

        wsCourse['A1'] = "Short Name"
        wsCourse['A2'] = "Full Name"
        wsCourse['A3'] = "Start Date"
        wsCourse['A4'] = "Backup Location"

        wsCourse['B1'] = objCourse.shortName
        wsCourse['B2'] = objCourse.fullName
        wsCourse['B3'] = formatTime(objCourse.startDate, '%Y-%m-%d')
        wsCourse['B4'] = objCourse.location

        setCellWidth(wsCourse)
        # writeSheet(wb, "Course", 0)

        return 1

    def sectionSheet(wb):
        wsSection = wb.create_sheet("Sections")

        wsSection['A1'] = "Week"
        wsSection['A2'] = "Name"
        wsSection['A3'] = "Summary"
        wsSection['A4'] = "Location"
        wsSection['A5'] = "Activities"

        for c in range(2, len(objSection) + 2):
            wsSection.cell(row=1, column=c).value = objSection[c - 2].number
            wsSection.cell(row=2, column=c).value = objSection[c - 2].name
            wsSection.cell(row=3, column=c).value = objSection[c - 2].summary
            wsSection.cell(row=4, column=c).value = objSection[c - 2].location
            wsSection.cell(row=5, column=c).value = objSection[c - 2].sequence

        setCellWidth(wsSection)

        return 1

    def activitySheet(wb):
        wsActivity = []
        y = 0
        for x in range(3, len(objSection) + 3):
            wsActivity.append(wb.create_sheet("Section" + str(x - 3)))

            wsActivity[x - 3].cell(row=1, column=1).value = "Module"
            wsActivity[x - 3].cell(row=2, column=1).value = "Section"
            wsActivity[x - 3].cell(row=3, column=1).value = "Section ID"
            wsActivity[x - 3].cell(row=4, column=1).value = "Module ID"
            wsActivity[x - 3].cell(row=5, column=1).value = "Activity ID"
            wsActivity[x - 3].cell(row=6, column=1).value = "Activity Location"
            wsActivity[x - 3].cell(row=7, column=1).value = "Module Location"
            wsActivity[x - 3].cell(row=8, column=1).value = "Name"
            wsActivity[x - 3].cell(row=9, column=1).value = "Intro"
            wsActivity[x - 3].cell(row=10, column=1).value = "Content"
            wsActivity[x - 3].cell(row=11, column=1).value = "URL"
            wsActivity[x - 3].cell(row=12, column=1).value = "Grade"
            wsActivity[x - 3].cell(row=13, column=1).value = "Due Date"

            for c in range(2, len(objSection[x - 3].activities) + 2):
                print "Module #: " + str(y)
                try:
                    wsActivity[x - 3].cell(row=1, column=c).value = objModule[y].modulename
                    wsActivity[x - 3].cell(row=2, column=c).value = objModule[y].sectionNumber
                    wsActivity[x - 3].cell(row=3, column=c).value = objModule[y].sectionID
                    wsActivity[x - 3].cell(row=4, column=c).value = objModule[y].moduleID
                    wsActivity[x - 3].cell(row=5, column=c).value = objModule[y].activityID
                    wsActivity[x - 3].cell(row=6, column=c).value = objModule[y].location
                    wsActivity[x - 3].cell(row=7, column=c).value = objModule[y].modulePath
                    wsActivity[x - 3].cell(row=8, column=c).value = objModule[y].name
                    wsActivity[x - 3].cell(row=9, column=c).value = objModule[y].intro
                    wsActivity[x - 3].cell(row=10, column=c).value = objModule[y].content
                    wsActivity[x - 3].cell(row=11, column=c).value = objModule[y].url
                    wsActivity[x - 3].cell(row=12, column=c).value = objModule[y].grade
                    wsActivity[x - 3].cell(row=13, column=c).value = formatTime(objModule[y].dueDate, '%Y-%m-%d %H:%M')
                except:
                    print "Failed on " + str(y)
                    pass
                y += 1
            setCellWidth(wsActivity[x - 3])
        y = None
        return 1

    def setCellWidth(ws):
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    try:
                        dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
                    except:
                        pass
        for col, value in dims.items():
            ws.column_dimensions[col].width = value

    def writeSheet(wb, name, numSheet):
        ws = wb.create_sheet(name, numSheet)
        return 1

    def readSheets():
        print wb.get_sheet_names()
        return 1

    choice(option)
